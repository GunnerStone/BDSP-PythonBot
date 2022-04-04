import openpyxl

# source: https://stackoverflow.com/questions/50491839/python-openpyxl-find-strings-in-column-and-return-row-number
def search_correct_trainer(sheet, trainer_type, trainer, pokemon):
    for row in range(1, sheet.max_row+1):
        trainer_name = "{}{}".format("B", row)
        type = "{}{}".format("C", row)
        pokemon1 = "{}{}".format("G", row)
        pokemon2 = "{}{}".format("H", row)
        pokemon3 = "{}{}".format("I", row)
        # if the trainer name is found in the column, check if the pokemon is in that row. If so, return that row.
        if sheet[trainer_name].value == trainer and sheet[type].value == trainer_type and ((pokemon in sheet[pokemon1].value) or (pokemon in sheet[pokemon2].value) or (pokemon in sheet[pokemon3].value)):
            return int(row)
    return None

def get_pokemon_info(sheet, pokemon_name):
    for row in range(1, sheet.max_row+1):
        row = int(row)
        coordinate = "{}{}".format("B", row)
        if sheet[coordinate].value == pokemon_name:
            data = {
                "Species": sheet.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("C")).value,
                "Nature": sheet.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("D")).value,
                "Item": sheet.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("E")).value,
                "Move1": sheet.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("F")).value,
                "Move2": sheet.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("G")).value,
                "Move3": sheet.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("H")).value,
                "Move4": sheet.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("I")).value,
                "Ability": sheet.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("J")).value,
                "Types": sheet.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("AH")).value,
                "HP": sheet.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("AB")).value,
                "Atk": sheet.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("AC")).value,
                "Def": sheet.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("AD")).value,
                "SpA": sheet.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("AE")).value,
                "SpD": sheet.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("AF")).value,
                "Spe": sheet.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("AG")).value
            }
            return data
    return None

# might want to add tower type in a later version to not take up so much memory haha...
# takes in trainer type, trainer name, and first pokemon as strings.
def get_trainer_info(trainer_type, trainer_name, first_pokemon):
    wb = openpyxl.load_workbook("battle_tower_info.xlsx")
    trainers = wb["Tower Trainers"]
    pokemon = wb["Tower Pokemon"]
    row = search_correct_trainer(trainers, trainer_type, trainer_name, first_pokemon)
    if row is None: return "Error: Trainer not found."
    pokemon1 = trainers.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("G")).value
    pokemon2 = trainers.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("H")).value
    pokemon3 = trainers.cell(row=row, column=openpyxl.utils.cell.column_index_from_string("I")).value
    pokemon1_data = get_pokemon_info(pokemon, pokemon1)
    pokemon2_data = get_pokemon_info(pokemon, pokemon2)
    pokemon3_data = get_pokemon_info(pokemon, pokemon3)
    return pokemon1_data, pokemon2_data, pokemon3_data